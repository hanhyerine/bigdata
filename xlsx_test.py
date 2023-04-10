from openpyxl import Workbook

# 엑셀파일 쓰기
write_wb = Workbook()

# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('생성시트')

# Sheet1에다 입력
write_ws = write_wb.active
write_ws['A1'] = '숫자'
write_ws['A2'] = '제목'
write_ws['A3'] = '평점'
write_ws['A4'] = '예매율'
write_ws['A5'] = '개봉날짜'

#행 단위로 추가
write_ws.append([1,2,3])

#셀 단위로 추가
write_ws.cell(5, 5, '5행5열')
write_wb.save('숫자.xlsx')